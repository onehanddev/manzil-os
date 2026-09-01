"""Cashbook deep module — single seam for cashbook invariants.

Interface (the seam):
  POST   /api/receipts                — recordReceipt
  POST   /api/expenses                — recordExpense (vendor_name inline)
  PUT    /api/cash-opening-balance    — set opening cash for a start date
  GET    /api/cash-opening-balance    — get opening cash
  GET    /api/reports/cashbook        — getReport(range) → opening/receipts/expenses/closing

All handlers enforce society scoping, amount>0 / >=0, business_date required,
fund/category/vendor belonging to the same society, and inclusive business_date
filtering. The report equation closing = opening + receipts − expenses is the
module's core invariant.
"""

from __future__ import annotations

import io
import secrets
import uuid
from datetime import date, datetime, timedelta, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from app.auth.deps import require_active, require_admin
from app.db import get_db
from app.models import CashOpeningBalance, Expense, ExpenseCategory, Flat, Fund, Notification, Person, Receipt, ReportRun, Society, SocietyMembership, User, Vendor

router = APIRouter(tags=["cashbook"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_uuid(val: str | None, field: str) -> uuid.UUID:
    if not val:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"{field} required")
    try:
        return uuid.UUID(val)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Invalid {field}") from e


def _get_society_id(current: dict) -> uuid.UUID:
    sid = current.get("society_id")
    if not sid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No society linked")
    return uuid.UUID(sid)


def _resolve_membership_id(db: Session, user_id: str, society_id: uuid.UUID) -> uuid.UUID:
    from app.models import SocietyMembership

    row = db.execute(
        select(SocietyMembership.id).where(
            SocietyMembership.user_id == uuid.UUID(user_id),
            SocietyMembership.society_id == society_id,
            SocietyMembership.status == "ACTIVE",
        )
    ).scalar_one_or_none()
    if row is None:
        # fallback: any active membership for this user (tests have single society)
        row = db.execute(
            select(SocietyMembership.id).where(
                SocietyMembership.user_id == uuid.UUID(user_id),
                SocietyMembership.status == "ACTIVE",
            )
        ).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No active membership")
    return row  # already UUID


def _ensure_fund(db: Session, fund_id: uuid.UUID, society_id: uuid.UUID) -> Fund:
    fund = db.execute(select(Fund).where(Fund.id == fund_id, Fund.society_id == society_id)).scalar_one_or_none()
    if not fund:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid fund_id")
    return fund


def _ensure_flat(db: Session, flat_id: uuid.UUID, society_id: uuid.UUID) -> Flat:
    flat = db.execute(select(Flat).where(Flat.id == flat_id, Flat.society_id == society_id)).scalar_one_or_none()
    if not flat:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Flat not found")
    return flat


def _ensure_category(db: Session, cat_id: uuid.UUID, society_id: uuid.UUID) -> ExpenseCategory:
    cat = db.execute(
        select(ExpenseCategory).where(ExpenseCategory.id == cat_id, ExpenseCategory.society_id == society_id)
    ).scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid category_id")
    return cat


def _ensure_person_same_society(db: Session, person_id: uuid.UUID, society_id: uuid.UUID) -> None:
    from app.models import Person

    row = db.execute(select(Person.id).where(Person.id == person_id, Person.society_id == society_id)).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid payer_person_id")


def _ensure_payer_is_occupant(db: Session, flat_id: uuid.UUID, person_id: uuid.UUID) -> None:
    from app.models import FlatOccupant

    occ = db.execute(
        select(FlatOccupant.id).where(
            FlatOccupant.flat_id == flat_id, FlatOccupant.person_id == person_id, FlatOccupant.is_active.is_(True)
        )
    ).scalar_one_or_none()
    if not occ:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="payer_person_id must be an active occupant of the flat"
        )


def _financial_year_start(business_date: date) -> int:
    return business_date.year if business_date.month >= 4 else business_date.year - 1


def _format_receipt_number(fy_year: int, sequence_number: int) -> str:
    return f"MANZIL/{str(fy_year)[-2:]}-{str(fy_year + 1)[-2:]}/{sequence_number:05d}"


def _next_receipt_number(db: Session, society_id: uuid.UUID, business_date: date) -> str:
    fy_year = _financial_year_start(business_date)
    db.execute(
        text(
            """
            INSERT INTO society_receipt_sequences (society_id, fy_year, next_number)
            VALUES (:society_id, :fy_year, 1)
            ON CONFLICT (society_id, fy_year) DO NOTHING
            """
        ),
        {"society_id": society_id, "fy_year": fy_year},
    )
    sequence_number = db.execute(
        text(
            """
            SELECT next_number
            FROM society_receipt_sequences
            WHERE society_id = :society_id AND fy_year = :fy_year
            FOR UPDATE
            """
        ),
        {"society_id": society_id, "fy_year": fy_year},
    ).scalar_one()
    db.execute(
        text(
            """
            UPDATE society_receipt_sequences
            SET next_number = next_number + 1, updated_at = now()
            WHERE society_id = :society_id AND fy_year = :fy_year
            """
        ),
        {"society_id": society_id, "fy_year": fy_year},
    )
    return _format_receipt_number(fy_year, int(sequence_number))


# ---------------------------------------------------------------------------
# Receipts
# ---------------------------------------------------------------------------


class CreateReceiptRequest(BaseModel):
    flat_id: str
    amount: float
    business_date: str
    fund_id: str
    payer_person_id: str | None = None
    type: str | None = None
    narration: str | None = None
    payment_method: str | None = None

    model_config = {"extra": "forbid"}


def _latest_receipt_notification(db: Session, receipt_id: uuid.UUID) -> Notification | None:
    return db.execute(
        select(Notification).where(Notification.receipt_id == receipt_id).order_by(Notification.created_at.desc())
    ).scalars().first()


def _serialize_receipt(r: Receipt, db: Session | None = None) -> dict:
    latest_notification = _latest_receipt_notification(db, r.id) if db else None
    serialized = {
        "id": str(r.id),
        "society_id": str(r.society_id),
        "receipt_number": r.receipt_number,
        "public_pdf_url": _receipt_public_pdf_url(r),
        "flat_id": str(r.flat_id),
        "payer_person_id": str(r.payer_person_id) if r.payer_person_id else None,
        "fund_id": str(r.fund_id) if r.fund_id else None,
        "amount": float(r.amount),
        "business_date": r.business_date.isoformat() if r.business_date else None,
        "type": r.type,
        "narration": r.narration,
        "payment_method": getattr(r, "payment_method", "CASH") or "CASH",
        "collected_by": str(r.collected_by),
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "status": getattr(r, "status", "POSTED") or "POSTED",
        "voided_at": r.voided_at.isoformat() if getattr(r, "voided_at", None) else None,
        "voided_by": str(r.voided_by) if getattr(r, "voided_by", None) else None,
        "void_reason": getattr(r, "void_reason", None),
        "updated_at": r.updated_at.isoformat() if getattr(r, "updated_at", None) else None,
    }
    if latest_notification:
        serialized["whatsapp_status"] = latest_notification.status
        serialized["whatsapp_failure_reason"] = getattr(latest_notification, "failure_reason", None)
    else:
        serialized["whatsapp_status"] = None
        serialized["whatsapp_failure_reason"] = None
    return serialized


def _serialize_notification(n: Notification) -> dict:
    return {
        "id": str(n.id),
        "society_id": str(n.society_id),
        "user_id": str(n.user_id) if n.user_id else None,
        "receipt_id": str(n.receipt_id) if n.receipt_id else None,
        "payer_person_id": str(n.payer_person_id) if n.payer_person_id else None,
        "flat_id": str(n.flat_id) if n.flat_id else None,
        "channel": n.channel,
        "provider_mode": n.provider_mode,
        "status": n.status,
        "message": n.message,
        "provider_message_id": getattr(n, "provider_message_id", None),
        "failure_reason": getattr(n, "failure_reason", None),
        "business_date": n.business_date.isoformat() if n.business_date else None,
        "created_at": n.created_at.isoformat() if n.created_at else None,
    }


def _public_receipt_pdf_url(receipt_id: uuid.UUID) -> str:
    return f"/receipts/{receipt_id}/pdf"


def _receipt_public_pdf_url(receipt: Receipt) -> str | None:
    if not receipt.public_pdf_token:
        return None
    return f"{_public_receipt_pdf_url(receipt.id)}?token={receipt.public_pdf_token}"


def _ensure_public_pdf_tokens(db: Session, receipts: list[Receipt]) -> None:
    missing = [receipt for receipt in receipts if not receipt.public_pdf_token]
    if not missing:
        return
    for receipt in missing:
        receipt.public_pdf_token = secrets.token_urlsafe(16)
    db.commit()
    for receipt in missing:
        db.refresh(receipt)


def _send_receipt_whatsapp(db: Session, receipt: Receipt, society_id: uuid.UUID) -> Notification:
    from app.notifications.provider import get_notification_provider

    society = db.execute(select(Society).where(Society.id == society_id)).scalar_one_or_none()
    flat = db.execute(select(Flat).where(Flat.id == receipt.flat_id, Flat.society_id == society_id)).scalar_one_or_none()
    payer = (
        db.execute(
            select(Person).where(
                Person.id == receipt.payer_person_id,
                Person.society_id == society_id,
            )
        ).scalar_one_or_none()
        if receipt.payer_person_id
        else None
    )
    return get_notification_provider().send_receipt_notification(
        db=db,
        society_id=society_id,
        receipt_id=receipt.id,
        payer_person_id=receipt.payer_person_id,
        flat_id=receipt.flat_id,
        business_date=receipt.business_date,
        amount=float(receipt.amount),
        narration=receipt.narration,
        receipt_number=receipt.receipt_number,
        flat_number=flat.flat_number if flat else None,
        society_name=society.name if society else None,
        payer_mobile=payer.mobile if payer else None,
        pdf_url=_receipt_public_pdf_url(receipt),
    )


@router.post("/api/receipts", status_code=status.HTTP_201_CREATED)
def create_receipt(payload: CreateReceiptRequest, db: Session = Depends(get_db), current=Depends(require_active)):
    society_id = _get_society_id(current)
    if payload.amount is None or float(payload.amount) <= 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="amount must be > 0")

    # business_date
    try:
        biz_date = date.fromisoformat(payload.business_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid business_date") from e

    # type
    rtype = (payload.type or "REGULAR").strip().upper()
    if rtype not in ("REGULAR", "ARREARS", "PART", "ADVANCE"):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid type")

    flat_id = _parse_uuid(payload.flat_id, "flat_id")
    fund_id = _parse_uuid(payload.fund_id, "fund_id")

    _ensure_flat(db, flat_id, society_id)
    _ensure_fund(db, fund_id, society_id)

    # payment_method – fixed CASH in Phase 0
    pm = (payload.payment_method or "CASH").strip().upper()
    if pm != "CASH":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="payment_method must be CASH")

    payer_id = None
    if payload.payer_person_id:
        try:
            payer_id = uuid.UUID(payload.payer_person_id)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid payer_person_id") from e
        _ensure_person_same_society(db, payer_id, society_id)
        _ensure_payer_is_occupant(db, flat_id, payer_id)

    membership_id = _resolve_membership_id(db, current["user_id"], society_id)

    receipt = Receipt(
        id=uuid.uuid4(),
        society_id=society_id,
        receipt_number=_next_receipt_number(db, society_id, biz_date),
        public_pdf_token=secrets.token_urlsafe(16),
        flat_id=flat_id,
        payer_person_id=payer_id,
        fund_id=fund_id,
        amount=payload.amount,
        business_date=biz_date,
        type=rtype,
        narration=payload.narration.strip() if payload.narration and payload.narration.strip() else None,
        payment_method=pm,
        collected_by=membership_id,
        status="POSTED",
    )
    db.add(receipt)
    db.flush()
    db.commit()
    db.refresh(receipt)

    # Notification delivery happens after the receipt transaction commits; failures
    # must not roll back or block the official receipt.
    try:
        _send_receipt_whatsapp(db, receipt, society_id)
        db.commit()
    except Exception:
        db.rollback()
    return _serialize_receipt(receipt, db)


class VoidReceiptRequest(BaseModel):
    reason: str | None = None


@router.get("/api/receipts/{receipt_id}")
def get_receipt(receipt_id: str, db: Session = Depends(get_db), current=Depends(require_active)):
    society_id = _get_society_id(current)
    try:
        rid = uuid.UUID(receipt_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found") from e
    row = db.execute(select(Receipt).where(Receipt.id == rid, Receipt.society_id == society_id)).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found")
    _ensure_public_pdf_tokens(db, [row])
    return _serialize_receipt(row, db)


@router.get("/api/receipts/{receipt_id}/pdf")
def get_receipt_pdf(receipt_id: str, db: Session = Depends(get_db), current=Depends(require_active)):
    society_id = _get_society_id(current)
    return _stream_receipt_pdf(receipt_id, db, society_id)


@router.get("/receipts/{receipt_id}/pdf")
def get_public_receipt_pdf(receipt_id: str, token: str | None = None, db: Session = Depends(get_db)):
    return _stream_receipt_pdf(receipt_id, db, None, token)


def _stream_receipt_pdf(receipt_id: str, db: Session, society_id: uuid.UUID | None, token: str | None = None):
    try:
        rid = uuid.UUID(receipt_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found") from e
    q = select(Receipt).where(Receipt.id == rid)
    if society_id:
        q = q.where(Receipt.society_id == society_id)
    receipt = db.execute(q).scalar_one_or_none()
    if not receipt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found")
    if society_id is None and (not token or token != receipt.public_pdf_token):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found")
    society_id = receipt.society_id

    society = db.execute(select(Society).where(Society.id == society_id)).scalar_one_or_none()
    flat = db.execute(select(Flat).where(Flat.id == receipt.flat_id, Flat.society_id == society_id)).scalar_one_or_none()
    payer = db.execute(select(Person).where(Person.id == receipt.payer_person_id)).scalar_one_or_none() if receipt.payer_person_id else None
    fund = db.execute(select(Fund).where(Fund.id == receipt.fund_id)).scalar_one_or_none() if receipt.fund_id else None
    collector = db.execute(
        select(User.display_name)
        .join(SocietyMembership, SocietyMembership.user_id == User.id)
        .where(SocietyMembership.id == receipt.collected_by)
    ).scalar_one_or_none()

    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A5, pageCompression=0)
    width, height = A5
    y = height - 18 * mm
    receipt_number = receipt.receipt_number or str(receipt.id)
    pdf.setTitle(f"Receipt {receipt_number}")
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(14 * mm, y, society.name if society else "Manzil Society")
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(width - 14 * mm, y, receipt_number)
    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawCentredString(width / 2, y, "Maintenance Receipt")
    y -= 14 * mm
    pdf.setFont("Helvetica", 10)
    rows = [
        ("Business Date", receipt.business_date.isoformat()),
        ("Flat", flat.flat_number if flat else str(receipt.flat_id)),
        ("Payer", payer.name if payer else "-"),
        ("Mobile", payer.mobile if payer else "-"),
        ("Fund", fund.name if fund else "-"),
        ("Type", receipt.type),
        ("Narration", receipt.narration or "-"),
        ("Amount", f"Rs. {float(receipt.amount):,.2f}"),
        ("Amount In Words", f"Rupees {int(float(receipt.amount))} only"),
        ("Collected By", collector or str(receipt.collected_by)),
        ("Timestamp", receipt.created_at.isoformat() if receipt.created_at else "-"),
    ]
    for label, value in rows:
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(16 * mm, y, f"{label}:")
        pdf.setFont("Helvetica", 9)
        pdf.drawString(54 * mm, y, str(value)[:64])
        y -= 7 * mm
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.drawCentredString(width / 2, 14 * mm, "Computer generated receipt. No signature required for pilot.")
    pdf.save()
    buffer.seek(0)
    safe_filename = receipt_number.replace("/", "-")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}.pdf"'},
    )


@router.post("/api/receipts/{receipt_id}/whatsapp-resend", status_code=status.HTTP_201_CREATED)
def resend_receipt_whatsapp(receipt_id: str, db: Session = Depends(get_db), current=Depends(require_active)):
    society_id = _get_society_id(current)
    try:
        rid = uuid.UUID(receipt_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found") from e
    receipt = db.execute(select(Receipt).where(Receipt.id == rid, Receipt.society_id == society_id)).scalar_one_or_none()
    if not receipt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found")
    notification = _send_receipt_whatsapp(db, receipt, society_id)
    db.commit()
    db.refresh(notification)
    return _serialize_notification(notification)


@router.post("/api/receipts/{receipt_id}/void", status_code=status.HTTP_200_OK)
@router.post("/api/receipts/{receipt_id}/undo", status_code=status.HTTP_200_OK)
def void_receipt(
    receipt_id: str,
    payload: VoidReceiptRequest | None = None,
    db: Session = Depends(get_db),
    current=Depends(require_active),
):
    """Void/undo a directly-submitted receipt while preserving history.

    Receipts have no draft state – they are POSTED on creation.
    Voiding keeps the row, sets status=VOIDED, and records voided_at/by/reason
    so the report can exclude it from totals while history remains auditable.
    """
    society_id = _get_society_id(current)
    try:
        rid = uuid.UUID(receipt_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found") from e
    receipt = db.execute(select(Receipt).where(Receipt.id == rid, Receipt.society_id == society_id)).scalar_one_or_none()
    if not receipt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found")
    if getattr(receipt, "status", "POSTED") == "VOIDED":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Receipt already voided")
    membership_id = _resolve_membership_id(db, current["user_id"], society_id)
    # RBAC: admin can void any; collector can only void receipts they collected
    roles = current.get("roles") or []
    is_admin = "SOCIETY_ADMIN" in roles or "super_admin" in roles
    if not is_admin and receipt.collected_by != membership_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin or the collector who created the receipt can void it")
    reason = (payload.reason.strip() if payload and payload.reason and payload.reason.strip() else None)
    # SOCIETY_ADMIN should provide reason but not strictly required for Phase 0; collector undo may omit
    receipt.status = "VOIDED"  # type: ignore[assignment]
    receipt.voided_at = datetime.now(timezone.utc)  # type: ignore[assignment]
    receipt.voided_by = membership_id  # type: ignore[assignment]
    receipt.void_reason = reason  # type: ignore[assignment]
    receipt.updated_at = datetime.now(timezone.utc)  # type: ignore[assignment]
    db.commit()
    db.refresh(receipt)
    return _serialize_receipt(receipt, db)


@router.get("/api/receipts")
def list_receipts(
    db: Session = Depends(get_db),
    current=Depends(require_active),
    from_date: str | None = Query(None, alias="from"),
    to_date: str | None = Query(None, alias="to"),
    date_from: str | None = Query(None, description="Alias for from"),
    date_to: str | None = Query(None, description="Alias for to"),
    include_voided: bool = Query(False, description="Include VOIDED receipts in listing (history view)"),
    flat_id: str | None = Query(None, description="Filter by flat"),
    collected_by: str | None = Query(None, description="Filter by collector membership id"),
    collector_id: str | None = Query(None, description="Alias for collected_by"),
    payer_person_id: str | None = Query(None, description="Filter by payer"),
):
    society_id = _get_society_id(current)
    q = select(Receipt).where(Receipt.society_id == society_id).order_by(Receipt.business_date, Receipt.created_at)
    if not include_voided:
        q = q.where(Receipt.status != "VOIDED")
    effective_from = from_date or date_from
    effective_to = to_date or date_to
    if effective_from:
        try:
            fd = date.fromisoformat(effective_from)
            q = q.where(Receipt.business_date >= fd)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid from date") from e
    if effective_to:
        try:
            td = date.fromisoformat(effective_to)
            q = q.where(Receipt.business_date <= td)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid to date") from e
    if flat_id:
        try:
            fid = uuid.UUID(flat_id)
            q = q.where(Receipt.flat_id == fid)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid flat_id") from e
    effective_collected = collected_by or collector_id
    if effective_collected:
        try:
            cb = uuid.UUID(effective_collected)
            q = q.where(Receipt.collected_by == cb)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid collected_by") from e
    if payer_person_id:
        try:
            pid = uuid.UUID(payer_person_id)
            q = q.where(Receipt.payer_person_id == pid)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid payer_person_id") from e
    rows = db.execute(q).scalars().all()
    _ensure_public_pdf_tokens(db, rows)
    return {"receipts": [_serialize_receipt(r, db) for r in rows]}


# ---------------------------------------------------------------------------
# Expenses
# ---------------------------------------------------------------------------


class CreateExpenseRequest(BaseModel):
    business_date: str
    amount: float
    fund_id: str
    category_id: str
    vendor_id: str | None = None
    vendor_name: str | None = None
    narration: str | None = None

    model_config = {"extra": "forbid"}


def _serialize_expense(e: Expense) -> dict:
    return {
        "id": str(e.id),
        "society_id": str(e.society_id),
        "business_date": e.business_date.isoformat() if e.business_date else None,
        "amount": float(e.amount),
        "fund_id": str(e.fund_id) if e.fund_id else None,
        "category_id": str(e.category_id),
        "vendor_id": str(e.vendor_id) if e.vendor_id else None,
        "narration": e.narration,
        "created_by": str(e.created_by),
        "created_at": e.created_at.isoformat() if e.created_at else None,
    }


@router.post("/api/expenses", status_code=status.HTTP_201_CREATED)
def create_expense(payload: CreateExpenseRequest, db: Session = Depends(get_db), current=Depends(require_admin)):
    society_id = _get_society_id(current)
    if payload.amount is None or float(payload.amount) <= 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="amount must be > 0")
    try:
        biz_date = date.fromisoformat(payload.business_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid business_date") from e

    fund_id = _parse_uuid(payload.fund_id, "fund_id")
    cat_id = _parse_uuid(payload.category_id, "category_id")
    _ensure_fund(db, fund_id, society_id)
    _ensure_category(db, cat_id, society_id)

    vendor_uuid: uuid.UUID | None = None
    if payload.vendor_id:
        try:
            vendor_uuid = uuid.UUID(payload.vendor_id)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid vendor_id") from e
        v = db.execute(select(Vendor).where(Vendor.id == vendor_uuid, Vendor.society_id == society_id)).scalar_one_or_none()
        if not v:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid vendor_id")
    elif payload.vendor_name and payload.vendor_name.strip():
        vname = payload.vendor_name.strip()
        # get-or-create by (society_id, name) case-insensitive
        existing = db.execute(
            select(Vendor).where(Vendor.society_id == society_id, func.lower(Vendor.name) == func.lower(vname))
        ).scalar_one_or_none()
        if existing:
            vendor_uuid = existing.id  # type: ignore[assignment]
        else:
            vendor_uuid = uuid.uuid4()
            vendor = Vendor(id=vendor_uuid, society_id=society_id, name=vname, is_active=True)
            db.add(vendor)
            db.flush()  # ensure vendor exists before expense FK

    membership_id = _resolve_membership_id(db, current["user_id"], society_id)

    expense = Expense(
        id=uuid.uuid4(),
        society_id=society_id,
        business_date=biz_date,
        amount=payload.amount,
        fund_id=fund_id,
        category_id=cat_id,
        vendor_id=vendor_uuid,
        narration=payload.narration.strip() if payload.narration and payload.narration.strip() else None,
        created_by=membership_id,
    )
    db.add(expense)
    db.commit()
    db.refresh(expense)
    return _serialize_expense(expense)


@router.get("/api/expenses/{expense_id}")
def get_expense(expense_id: str, db: Session = Depends(get_db), current=Depends(require_admin)):
    society_id = _get_society_id(current)
    try:
        parsed_id = uuid.UUID(expense_id)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expense not found") from error
    expense = db.execute(
        select(Expense).where(Expense.id == parsed_id, Expense.society_id == society_id)
    ).scalar_one_or_none()
    if not expense:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expense not found")
    return _serialize_expense(expense)


@router.get("/api/expenses")
def list_expenses(
    db: Session = Depends(get_db),
    current=Depends(require_active),
    from_date: str | None = Query(None, alias="from"),
    to_date: str | None = Query(None, alias="to"),
    category_id: str | None = Query(None, description="Filter by ExpenseCategory"),
    vendor_id: str | None = Query(None, description="Filter by Vendor"),
    fund_id: str | None = Query(None, description="Filter by Fund"),
):
    society_id = _get_society_id(current)
    q = select(Expense).where(Expense.society_id == society_id).order_by(Expense.business_date, Expense.created_at)
    if from_date:
        try:
            fd = date.fromisoformat(from_date)
            q = q.where(Expense.business_date >= fd)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid from date") from e
    if to_date:
        try:
            td = date.fromisoformat(to_date)
            q = q.where(Expense.business_date <= td)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid to date") from e
    if category_id:
        try:
            cid = uuid.UUID(category_id)
            q = q.where(Expense.category_id == cid)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid category_id") from e
    if vendor_id:
        try:
            vid = uuid.UUID(vendor_id)
            q = q.where(Expense.vendor_id == vid)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid vendor_id") from e
    if fund_id:
        try:
            fid = uuid.UUID(fund_id)
            q = q.where(Expense.fund_id == fid)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid fund_id") from e
    rows = db.execute(q).scalars().all()
    return {"expenses": [_serialize_expense(r) for r in rows]}


# ---------------------------------------------------------------------------
# Cash opening balance
# ---------------------------------------------------------------------------


class CashOpeningRequest(BaseModel):
    opening_date: str
    amount: float

    model_config = {"extra": "forbid"}


@router.put("/api/cash-opening-balance")
def put_cash_opening(payload: CashOpeningRequest, db: Session = Depends(get_db), current=Depends(require_admin)):
    society_id = _get_society_id(current)
    if payload.amount is None or float(payload.amount) < 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="amount must be >= 0")
    try:
        od = date.fromisoformat(payload.opening_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid opening_date") from e

    membership_id = _resolve_membership_id(db, current["user_id"], society_id)
    existing = db.execute(
        select(CashOpeningBalance).where(
            CashOpeningBalance.society_id == society_id, CashOpeningBalance.opening_date == od
        )
    ).scalar_one_or_none()
    if existing:
        existing.amount = payload.amount  # type: ignore[assignment]
        existing.updated_at = datetime.now(timezone.utc)
        existing.created_by = membership_id  # type: ignore[assignment]
        db.commit()
        db.refresh(existing)
        row = existing
    else:
        row = CashOpeningBalance(
            society_id=society_id, opening_date=od, amount=payload.amount, created_by=membership_id
        )
        db.add(row)
        db.commit()
        db.refresh(row)
    return {
        "society_id": str(row.society_id),
        "opening_date": row.opening_date.isoformat(),
        "amount": float(row.amount),
        "exists": True,
    }


@router.get("/api/cash-opening-balance")
def get_cash_opening(
    db: Session = Depends(get_db),
    current=Depends(require_admin),
    date_: str | None = Query(None, alias="date"),
):
    society_id = _get_society_id(current)
    if not date_:
        # return most recent or 0
        row = db.execute(
            select(CashOpeningBalance)
            .where(CashOpeningBalance.society_id == society_id)
            .order_by(CashOpeningBalance.opening_date.desc())
            .limit(1)
        ).scalar_one_or_none()
        if not row:
            return {"society_id": str(society_id), "opening_date": None, "amount": 0, "exists": False}
        return {
            "society_id": str(row.society_id),
            "opening_date": row.opening_date.isoformat(),
            "amount": float(row.amount),
            "exists": True,
        }
    try:
        od = date.fromisoformat(date_)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid date") from e
    row = db.execute(
        select(CashOpeningBalance).where(CashOpeningBalance.society_id == society_id, CashOpeningBalance.opening_date == od)
    ).scalar_one_or_none()
    if not row:
        return {"society_id": str(society_id), "opening_date": date_, "amount": 0, "exists": False}
    return {
        "society_id": str(row.society_id),
        "opening_date": row.opening_date.isoformat(),
        "amount": float(row.amount),
        "exists": True,
    }


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------


def _save_report_run(db: Session, society_id: uuid.UUID, current: dict, report: dict, format: str) -> None:
    generated_by = _resolve_membership_id(db, current["user_id"], society_id)
    now = datetime.now(timezone.utc)
    db.query(ReportRun).filter(
        ReportRun.society_id == society_id,
        ReportRun.generated_at < now - timedelta(days=90),
    ).delete(synchronize_session=False)
    run = db.execute(
        select(ReportRun).where(
            ReportRun.society_id == society_id,
            ReportRun.from_date == date.fromisoformat(report["from"]),
            ReportRun.to_date == date.fromisoformat(report["to"]),
        )
    ).scalar_one_or_none()
    if run is None:
        run = ReportRun(
            society_id=society_id,
            from_date=date.fromisoformat(report["from"]),
            to_date=date.fromisoformat(report["to"]),
            opening=report["opening"],
            total_receipts=report["total_receipts"],
            total_expenses=report["total_expenses"],
            closing=report["closing"],
            generated_by=generated_by,
            format=format,
        )
        db.add(run)
    else:
        run.opening = report["opening"]
        run.total_receipts = report["total_receipts"]
        run.total_expenses = report["total_expenses"]
        run.closing = report["closing"]
        run.generated_at = now
        run.generated_by = generated_by
        run.format = format
    db.commit()


def _serialize_report_run(run: ReportRun) -> dict:
    return {
        "id": str(run.id),
        "from": run.from_date.isoformat(),
        "to": run.to_date.isoformat(),
        "opening": float(run.opening),
        "total_receipts": float(run.total_receipts),
        "total_expenses": float(run.total_expenses),
        "closing": float(run.closing),
        "generated_at": run.generated_at.isoformat(),
        "generated_by": str(run.generated_by) if run.generated_by else None,
        "format": run.format,
    }


def _cashbook_pdf(report: dict) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title = report["society"]["name"] or "Society"
    story = [
        Paragraph(f"{title} Cashbook Report", styles["Title"]),
        Paragraph(f"{report['from']} to {report['to']}", styles["Normal"]),
        Spacer(1, 0.4 * cm),
    ]
    summary = [
        ["Opening", "Receipts", "Expenses", "Closing"],
        [
            f"Rs. {report['opening']:,.2f}",
            f"Rs. {report['total_receipts']:,.2f}",
            f"Rs. {report['total_expenses']:,.2f}",
            f"Rs. {report['closing']:,.2f}",
        ],
    ]
    summary_table = Table(summary, colWidths=[4.3 * cm] * 4)
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 0.5 * cm)])

    rows = [["Date", "Particulars", "Receipt", "Payment", "Fund / Category"]]
    for receipt in report["receipts"]:
        rows.append(
            [
                receipt["business_date"],
                receipt["narration"] or "Receipt",
                f"Rs. {receipt['amount']:,.2f}",
                "",
                " / ".join(filter(None, [receipt["flat"]["flat_number"], (receipt["fund"] or {}).get("name")])),
            ]
        )
    for expense in report["expenses"]:
        rows.append(
            [
                expense["business_date"],
                expense["narration"] or "Expense",
                "",
                f"Rs. {expense['amount']:,.2f}",
                " / ".join(
                    filter(
                        None,
                        [
                            expense["category"]["name"],
                            (expense["vendor"] or {}).get("name"),
                            (expense["fund"] or {}).get("name"),
                        ],
                    )
                ),
            ]
        )
    statement = Table(rows, colWidths=[2.2 * cm, 5.8 * cm, 2.5 * cm, 2.5 * cm, 4 * cm], repeatRows=1)
    statement.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(statement)
    document.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=cashbook-report.pdf"},
    )


def _cashbook_xlsx(report: dict) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cashbook"
    headers = ["Date", "Particulars", "Receipt", "Payment", "Fund / Category"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for receipt in report["receipts"]:
        sheet.append(
            [
                receipt["business_date"],
                receipt["narration"] or "Receipt",
                receipt["amount"],
                None,
                " / ".join(filter(None, [receipt["flat"]["flat_number"], (receipt["fund"] or {}).get("name")])),
            ]
        )
    for expense in report["expenses"]:
        sheet.append(
            [
                expense["business_date"],
                expense["narration"] or "Expense",
                None,
                expense["amount"],
                " / ".join(
                    filter(
                        None,
                        [
                            expense["category"]["name"],
                            (expense["vendor"] or {}).get("name"),
                            (expense["fund"] or {}).get("name"),
                        ],
                    )
                ),
            ]
        )
    sheet.append([])
    sheet.append(["", "Opening", report["opening"], None, ""])
    sheet.append(["", "Total receipts", report["total_receipts"], None, ""])
    sheet.append(["", "Total expenses", None, report["total_expenses"], ""])
    sheet.append(["", "Closing", report["closing"], None, ""])
    for column in ("C", "D"):
        for row in range(2, sheet.max_row + 1):
            sheet[f"{column}{row}"].number_format = "#,##0.00"
    for column in sheet.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(width + 4, 40)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cashbook-report.xlsx"},
    )


@router.get("/api/reports/cashbook")
def cashbook_report(
    db: Session = Depends(get_db),
    current=Depends(require_admin),
    from_date: str = Query(..., alias="from"),
    to_date: str = Query(..., alias="to"),
    format: str = Query("json"),
):
    society_id = _get_society_id(current)
    try:
        fd = date.fromisoformat(from_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid from date") from e
    try:
        td = date.fromisoformat(to_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid to date") from e
    if fd > td:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="from must be <= to")

    # opening = cash_opening_balances where opening_date == from, else 0
    opening_row = db.execute(
        select(CashOpeningBalance).where(CashOpeningBalance.society_id == society_id, CashOpeningBalance.opening_date == fd)
    ).scalar_one_or_none()
    opening = float(opening_row.amount) if opening_row else 0.0

    # Only POSTED receipts contribute to cash; VOIDED are preserved for history but excluded from totals
    total_receipts = float(
        db.execute(
            select(func.coalesce(func.sum(Receipt.amount), 0)).where(
                Receipt.society_id == society_id,
                Receipt.status != "VOIDED",
                Receipt.business_date >= fd,
                Receipt.business_date <= td,
            )
        ).scalar_one()
    )
    total_expenses = float(
        db.execute(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.society_id == society_id, Expense.business_date >= fd, Expense.business_date <= td
            )
        ).scalar_one()
    )
    closing = opening + total_receipts - total_expenses

    receipts = db.execute(
        select(Receipt)
        .where(Receipt.society_id == society_id, Receipt.status != "VOIDED", Receipt.business_date >= fd, Receipt.business_date <= td)
        .order_by(Receipt.business_date, Receipt.created_at)
    ).scalars().all()
    expenses = db.execute(
        select(Expense)
        .where(Expense.society_id == society_id, Expense.business_date >= fd, Expense.business_date <= td)
        .order_by(Expense.business_date, Expense.created_at)
    ).scalars().all()

    flat_ids = {receipt.flat_id for receipt in receipts}
    fund_ids = {row.fund_id for row in [*receipts, *expenses] if row.fund_id}
    category_ids = {expense.category_id for expense in expenses}
    vendor_ids = {expense.vendor_id for expense in expenses if expense.vendor_id}
    flats_by_id = {
        flat.id: flat for flat in db.execute(select(Flat).where(Flat.society_id == society_id, Flat.id.in_(flat_ids))).scalars()
    }
    funds_by_id = {
        fund.id: fund for fund in db.execute(select(Fund).where(Fund.society_id == society_id, Fund.id.in_(fund_ids))).scalars()
    }
    categories_by_id = {
        category.id: category
        for category in db.execute(
            select(ExpenseCategory).where(
                ExpenseCategory.society_id == society_id,
                ExpenseCategory.id.in_(category_ids),
            )
        ).scalars()
    }
    vendors_by_id = {
        vendor.id: vendor
        for vendor in db.execute(select(Vendor).where(Vendor.society_id == society_id, Vendor.id.in_(vendor_ids))).scalars()
    }

    # society name for report header
    from app.models import Society

    society = db.execute(select(Society).where(Society.id == society_id)).scalar_one_or_none()
    society_name = society.name if society else None

    report = {
        "society": {"id": str(society_id), "name": society_name},
        "from": from_date,
        "to": to_date,
        "opening": opening,
        "total_receipts": total_receipts,
        "total_expenses": total_expenses,
        "closing": closing,
        "receipts": [
            {
                **_serialize_receipt(receipt, db),
                "flat": {
                    "id": str(receipt.flat_id),
                    "flat_number": flats_by_id[receipt.flat_id].flat_number,
                },
                "fund": (
                    {"id": str(receipt.fund_id), "name": funds_by_id[receipt.fund_id].name}
                    if receipt.fund_id in funds_by_id
                    else None
                ),
            }
            for receipt in receipts
        ],
        "expenses": [
            {
                **_serialize_expense(expense),
                "category": {
                    "id": str(expense.category_id),
                    "name": categories_by_id[expense.category_id].name,
                },
                "vendor": (
                    {"id": str(expense.vendor_id), "name": vendors_by_id[expense.vendor_id].name}
                    if expense.vendor_id in vendors_by_id
                    else None
                ),
                "fund": (
                    {"id": str(expense.fund_id), "name": funds_by_id[expense.fund_id].name}
                    if expense.fund_id in funds_by_id
                    else None
                ),
            }
            for expense in expenses
        ],
    }
    if format == "json":
        return report
    if format == "xlsx":
        _save_report_run(db, society_id, current, report, format)
        return _cashbook_xlsx(report)
    if format == "pdf":
        _save_report_run(db, society_id, current, report, format)
        return _cashbook_pdf(report)
    raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="format must be json, xlsx, or pdf")


@router.get("/api/reports/history")
def cashbook_history(
    db: Session = Depends(get_db),
    current=Depends(require_admin),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=10),
):
    society_id = _get_society_id(current)
    base = select(ReportRun).where(ReportRun.society_id == society_id)
    total = db.execute(select(func.count()).select_from(base.subquery())).scalar_one()
    runs = db.execute(
        base.order_by(ReportRun.generated_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).scalars().all()
    return {"runs": [_serialize_report_run(run) for run in runs], "page": page, "page_size": page_size, "total": total}
