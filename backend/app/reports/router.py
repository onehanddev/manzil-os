"""Flat dues Excel export – society-scoped, derived dues."""

import io
import uuid

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.auth.deps import require_active
from app.db import get_db
from app.models import Flat, FlatCategory, OpeningDue, Receipt

router = APIRouter(tags=["reports"])


def _build_excel_response(db: Session, society_id: uuid.UUID, is_active: bool | None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    q = select(Flat).where(Flat.society_id == society_id)
    if is_active is not None:
        q = q.where(Flat.is_active.is_(is_active))
    q = q.order_by(Flat.flat_number)
    flats = db.execute(q).scalars().all()

    # bulk categories
    cat_ids = {f.flat_category_id for f in flats}
    cats = {}
    if cat_ids:
        cats = {c.id: c for c in db.execute(select(FlatCategory).where(FlatCategory.id.in_(cat_ids))).scalars().all()}

    flat_ids = [f.id for f in flats]
    # occupant maps for POC
    from app.flats.router import _build_occupant_maps

    occ_maps = _build_occupant_maps(db, flat_ids) if flat_ids else {}

    # opening map
    opening_map: dict[uuid.UUID, float] = {}
    if flat_ids:
        for od in db.execute(select(OpeningDue).where(OpeningDue.flat_id.in_(flat_ids))).scalars().all():
            opening_map[od.flat_id] = float(od.amount)

    # total paid map (POSTED only)
    paid_map: dict[uuid.UUID, float] = {}
    if flat_ids:
        rows = db.execute(
            select(Receipt.flat_id, func.coalesce(func.sum(Receipt.amount), 0)).where(
                Receipt.flat_id.in_(flat_ids), Receipt.status != "VOIDED", Receipt.society_id == society_id
            ).group_by(Receipt.flat_id)
        ).all()
        paid_map = {fid: float(total) for fid, total in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = "Flat Dues"
    headers = ["Flat", "Category", "POC Name", "POC Mobile", "Opening Due", "Total Paid", "Current Due"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for flat in flats:
        cat = cats.get(flat.flat_category_id)
        cat_name = cat.name if cat else ""
        occ = occ_maps.get(flat.id, {})
        poc = occ.get("tenant") or occ.get("owner")
        poc_name = poc["name"] if poc else ""
        poc_mobile = poc["mobile"] if poc else ""
        opening = opening_map.get(flat.id, 0.0)
        paid = paid_map.get(flat.id, 0.0)
        current = opening - paid
        ws.append([flat.flat_number, cat_name, poc_name, poc_mobile, opening, paid, current])

    # formatting
    for col in ["E", "F", "G"]:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = "#,##0.00"
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # freeze header
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/api/reports/flat-dues.xlsx")
def export_flat_dues_api(
    db: Session = Depends(get_db),
    current=Depends(require_active),
    is_active: bool | None = Query(None, description="Filter by is_active optionally"),
):
    sid = current.get("society_id")
    if not sid:
        from fastapi import HTTPException, status

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No society linked")
    society_id = uuid.UUID(sid)
    buf = _build_excel_response(db, society_id, is_active)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flat-dues.xlsx"},
    )


@router.get("/reports/flat-dues.xlsx")
def export_flat_dues_alias(
    db: Session = Depends(get_db),
    current=Depends(require_active),
    is_active: bool | None = Query(None),
):
    sid = current.get("society_id")
    if not sid:
        from fastapi import HTTPException, status

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No society linked")
    society_id = uuid.UUID(sid)
    buf = _build_excel_response(db, society_id, is_active)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flat-dues.xlsx"},
    )
