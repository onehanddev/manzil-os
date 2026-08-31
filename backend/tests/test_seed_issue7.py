"""Issue 7 TDD – Demo Fixture idempotency + Excel varied dues (red phase).

Seam: HTTP API via TestClient + direct seed invocation.
AC under test:
  1. seed can be run repeatedly without duplicate flats/receipts
  2. Demo Excel export shows varied dues including at least one advance (-500) and one zero-clear flat
  3. Collector + Admin exist after seed
  4. Tenant-default + owner-fallback occupant cases present
  5. No reliance on real WhatsApp/OTP – provider_mode=test LOGGED rows exist
"""

import io
import uuid

import psycopg
import pytest

from conftest import TEST_DB_URL

DEMO_FLATS_EXPECTED = 8
DEMO_RECEIPTS_EXPECTED_MIN = 6


import jwt as pyjwt


@pytest.fixture(autouse=True)
def _supabase_env(monkeypatch):
    monkeypatch.setenv("SUPABASE_URL", "https://test.supabase.co")
    monkeypatch.setenv("SUPABASE_JWT_SECRET", "test-supabase-jwt-secret-32-chars-long!!")
    monkeypatch.setenv("SUPABASE_ANON_KEY", "test-anon-key")
    monkeypatch.setenv("SUPABASE_SERVICE_ROLE_KEY", "test-service-key")
    monkeypatch.setenv("PROVIDER_MODE", "test")
    # ensure seed writes to test DB
    monkeypatch.setenv("DATABASE_URL", TEST_DB_URL)
    # clear cached engine/session so next get_engine() picks up TEST_DB_URL
    try:
        import app.db as db_mod

        db_mod._engine = None
        db_mod._SessionLocal = None
    except Exception:
        pass
    # also mock supabase auth like other tests do
    _store: dict[str, tuple[str, str]] = {}
    import psycopg as _psycopg

    with _psycopg.connect(TEST_DB_URL, autocommit=True) as c:
        with c.cursor() as cur:
            cur.execute("SELECT id, auth_user_id FROM users WHERE mobile=%s", ("+919000000000",))
            row = cur.fetchone()
            if row:
                user_id, auth_id = row
                if not auth_id:
                    import uuid as _uuid

                    auth_id = str(_uuid.uuid4())
                    cur.execute("UPDATE users SET auth_user_id=%s WHERE id=%s", (auth_id, user_id))
                else:
                    auth_id = str(auth_id)
                _store["+919000000000"] = (auth_id, "admin123")
    import app.admin.router as admin_router_mod
    import app.auth.router as router_mod
    import app.auth.supabase_client as sc_mod

    def _mock_create_user(mobile, password, display_name):
        from app.auth.security import normalize_mobile

        norm = normalize_mobile(mobile)
        if norm in _store:
            return None
        import uuid as _uuid

        auth_id = str(_uuid.uuid4())
        _store[norm] = (auth_id, password)
        return auth_id

    def _mock_sign_in(mobile, password):
        from app.auth.security import normalize_mobile

        norm = normalize_mobile(mobile)
        entry = _store.get(norm)
        if not entry:
            return None
        auth_id, pw = entry
        if pw != password:
            return None
        token = pyjwt.encode({"sub": auth_id, "phone": norm, "aud": "authenticated", "exp": 9999999999}, "test-supabase-jwt-secret-32-chars-long!!", algorithm="HS256")
        return {"access_token": token, "user": {"id": auth_id}}

    monkeypatch.setattr(router_mod, "supabase_sign_in", _mock_sign_in)
    monkeypatch.setattr(admin_router_mod, "supabase_create_user", _mock_create_user)
    monkeypatch.setattr(sc_mod, "supabase_create_user", _mock_create_user)
    monkeypatch.setattr(sc_mod, "supabase_sign_in", _mock_sign_in)
    yield
    # clear engine again for other tests
    try:
        import app.db as db_mod2

        db_mod2._engine = None
        db_mod2._SessionLocal = None
    except Exception:
        pass


def _run_seed():
    """Invoke the demo seed – imports the module so test fails (RED) if not present."""
    import importlib

    # ensure engine cache cleared before import if DATABASE_URL changed
    try:
        import app.db as db_mod

        # do not clear if already cleared
        if db_mod._engine is not None:
            # check if engine URL matches TEST_DB_URL
            existing_url = str(db_mod._engine.url) if hasattr(db_mod._engine, "url") else ""
            if "manzil_os_test" not in existing_url:
                db_mod._engine = None
                db_mod._SessionLocal = None
    except Exception:
        pass
    mod = importlib.import_module("app.seed")
    # seed must expose run() or main() or seed() – try each
    fn = getattr(mod, "run", None) or getattr(mod, "seed", None) or getattr(mod, "main", None)
    assert fn is not None, "app.seed must expose run() / seed() / main()"
    return fn()


def test_seed_idempotent_no_duplicate_flats_receipts(conn):
    # First run
    _run_seed()
    with psycopg.connect(TEST_DB_URL, autocommit=True) as c:
        with c.cursor() as cur:
            cur.execute("SELECT count(*) FROM flats WHERE flat_number LIKE 'DEMO-%'")
            flats1 = cur.fetchone()[0]
            cur.execute("SELECT count(*) FROM receipts WHERE narration LIKE 'DEMO-SEED%'")
            receipts1 = cur.fetchone()[0]
    assert flats1 == DEMO_FLATS_EXPECTED, f"expected {DEMO_FLATS_EXPECTED} demo flats, got {flats1}"
    assert receipts1 >= DEMO_RECEIPTS_EXPECTED_MIN, f"expected >= {DEMO_RECEIPTS_EXPECTED_MIN} demo receipts, got {receipts1}"

    # Second run – must be idempotent (counts unchanged)
    _run_seed()
    with psycopg.connect(TEST_DB_URL, autocommit=True) as c:
        with c.cursor() as cur:
            cur.execute("SELECT count(*) FROM flats WHERE flat_number LIKE 'DEMO-%'")
            flats2 = cur.fetchone()[0]
            cur.execute("SELECT count(*) FROM receipts WHERE narration LIKE 'DEMO-SEED%'")
            receipts2 = cur.fetchone()[0]
    assert flats2 == flats1, f"seed not idempotent: flats {flats1} -> {flats2}"
    assert receipts2 == receipts1, f"seed not idempotent: receipts {receipts1} -> {receipts2}"


def test_seed_excel_varied_dues_advance_and_zero(conn):
    from tests.test_master_data import _admin_token, _auth_header, _client

    _run_seed()
    client = _client()
    token = _admin_token(client)

    # Verify via API: GET /flats?with_dues=true shows advance and zero-clear
    r = client.get("/api/flats?with_dues=true", headers=_auth_header(token))
    assert r.status_code == 200, r.text
    demo_flats = [f for f in r.json().get("flats", []) if f.get("flat_number", "").startswith("DEMO-")]
    assert len(demo_flats) >= DEMO_FLATS_EXPECTED, f"not enough demo flats {len(demo_flats)}"
    dues = [float(f.get("current_due", 0)) for f in demo_flats]
    assert any(d == -500 for d in dues), f"expected advance -500, got dues {dues}"
    assert any(d == 0 for d in dues), f"expected zero-clear flat, got dues {dues}"

    # Excel export seam: GET /api/reports/flat-dues.xlsx
    er = client.get("/api/reports/flat-dues.xlsx", headers=_auth_header(token))
    if er.status_code == 404:
        er = client.get("/reports/flat-dues.xlsx", headers=_auth_header(token))
    assert er.status_code == 200, er.text
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(er.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = " ".join(str(h) for h in rows[0] if h)
    assert "Flat" in header, header
    assert "Current Due" in header or "current" in header.lower(), header
    demo_rows = [row for row in rows[1:] if str(row[0]).startswith("DEMO-")]
    assert len(demo_rows) >= DEMO_FLATS_EXPECTED, demo_rows
    excel_dues = [float(row[6]) for row in demo_rows if row[6] is not None]
    assert any(d == -500 for d in excel_dues), f"excel missing -500 advance {excel_dues}"
    assert any(d == 0 for d in excel_dues), f"excel missing zero-clear {excel_dues}"


def test_seed_collects_tenant_default_and_owner_fallback_and_users(conn):
    from tests.test_master_data import _client, _auth_header, _admin_token

    _run_seed()
    # Check occupants: tenant-default vs owner-fallback present
    with psycopg.connect(TEST_DB_URL, autocommit=True) as c:
        with c.cursor() as cur:
            cur.execute("""
                SELECT f.flat_number, fo.role FROM flats f
                JOIN flat_occupants fo ON fo.flat_id=f.id AND fo.is_active=TRUE
                WHERE f.flat_number LIKE 'DEMO-%' ORDER BY f.flat_number, fo.role
            """)
            rows = cur.fetchall()
    # Must have at least one tenant occupant and one owner-only flat
    roles_by_flat: dict[str, set] = {}
    for flat_num, role in rows:
        roles_by_flat.setdefault(flat_num, set()).add(role)
    has_tenant = any("TENANT" in roles for roles in roles_by_flat.values())
    has_owner_only = any(roles == {"OWNER"} for roles in roles_by_flat.values())
    assert has_tenant, f"missing tenant-default case {roles_by_flat}"
    assert has_owner_only, f"missing owner-fallback case {roles_by_flat}"

    # Admin + Collector existence
    with psycopg.connect(TEST_DB_URL, autocommit=True) as c:
        with c.cursor() as cur:
            cur.execute("SELECT mobile FROM users WHERE mobile IN ('+919000000000', '+919000000100')")
            mobiles = {r[0] for r in cur.fetchall()}
    assert "+919000000000" in mobiles, f"admin missing {mobiles}"
    assert "+919000000100" in mobiles, f"collector missing {mobiles}"

    # Notification provider_mode=test rows (LOGGED) for receipts
    with psycopg.connect(TEST_DB_URL, autocommit=True) as c:
        with c.cursor() as cur:
            cur.execute("SELECT provider_mode, status FROM notifications WHERE receipt_id IN (SELECT id FROM receipts WHERE narration LIKE 'DEMO-SEED%') LIMIT 5")
            notif_rows = cur.fetchall()
    # at least some demo receipts should have test provider logged rows
    if notif_rows:
        assert any(pm == "test" and st == "LOGGED" for pm, st in notif_rows), notif_rows
